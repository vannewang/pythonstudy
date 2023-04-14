# coding=utf-8
import os
# openpyxl version 3.0.5
import openpyxl
# canan version 0.5.19
import cn2an
from openpyxl.drawing.image import Image


# excel操作类。 将excel中某一列的部分单元格中的名称进行匹配，并将另外一个文件夹中名称匹配的图片插入进来。
# 匹配规则： excel单元格中的格式： 图二十一； 文件夹中的图片名称：21.jpg

def excel():
    # 需要处理的文件夹路径和 Excel 文件路径
    global img, chinese
    folder_path = "C:/Users/wangyu/Desktop/test/pic"
    excel_path = "C:/Users/wangyu/Desktop/test/excel.xlsx"

    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(excel_path)
    # 遍历 Excel 文件中的所有 sheet 页
    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        # 获取图片数字和对应行号的字典
        chinese_dict = {}
        for row in worksheet.iter_rows(min_row=2, max_col=1):
            text = row[0].value
            if text and text.startswith('图'):
                text = text.strip()
                text_split = text[1:]
                num = cn2an.cn2an(text_split)
                chinese_dict[num] = row[0].row
        set_cell_height_witth(chinese_dict, worksheet)
        # 保存 Excel 文件
        workbook.save(excel_path)
        add_image(chinese_dict, folder_path, worksheet)
        # 保存 Excel 文件
        workbook.save(excel_path)
        print("当前sheet: {0}处理完成!".format(sheet_name))


# 设置单元格大小
def set_cell_height_witth(chinese_dict, worksheet):
    if len(chinese_dict) > 0:
        for row in chinese_dict:
            # 设置高度
            num = chinese_dict[row]
            worksheet.row_dimensions[num].height = 200
    worksheet.column_dimensions['A'].width = 40


# 添加图片
def add_image(chinese_dict, folder_path, worksheet):
    global img
    # 遍历文件夹中的所有图片文件
    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            # 从文件名中提取数字
            num = int(filename.split(".")[0])
            # 根据图片数字查找对应的行号
            row_num = chinese_dict.get(num)
            if row_num is not None:
                chinese_cell = worksheet.cell(row=row_num, column=1)
                # 将图片路径整合为绝对路径
                img_path = os.path.join(folder_path, filename)
                # 在当前 sheet 页的对应单元格中插入图片
                img = Image(img_path)
                # 获取图片宽度和高度
                width, height = img.width, img.height
                cell_width = worksheet.column_dimensions["A"].width * 5
                cell_height = worksheet.row_dimensions[chinese_cell.row].height
                # 图片缩放比例
                img_scale = 1
                if width > height:
                    img_scale = height / width
                    img.width = (cell_height * img_scale) * 2
                    img.height = cell_height
                else:
                    img_scale = width / height
                    img.width = cell_width * 2
                    img.height = cell_width * img_scale
                # 添加图像到工作表
                anchor = f'{chinese_cell.column_letter}{chinese_cell.row}'
                worksheet.add_image(img, anchor)
